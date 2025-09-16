import os
import pandas as pd
from . import AmpuMapel, Berita, EvaluasiGuru, Kehadiran, Keterangan, PembagianKelas, Penilaian, TahunSemester, User, app, db, Kbm, Kelas, Siswa, Guru, Mapel, Semester, TahunAkademik, time_zone_wib
from flask import flash, redirect, request, jsonify, url_for, render_template, session, abort, send_file
from sqlalchemy import case, extract, func, literal
from sqlalchemy.exc import IntegrityError
from datetime import datetime
from io import BytesIO

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
@app.route("/kbm/list")
def list_kbm():
    if 'role' not in session or session['role'] == 'murid':
        return redirect(url_for('login'))
    if session['role'] == 'guru':
        guru = Guru.query.filter_by(nip=session['nip']).first()
        data_ampu = AmpuMapel.query.filter_by(nip=guru.nip).all()
        daftar_pembagian = PembagianKelas.query.filter_by(nip=guru.nip).all()
        query = AmpuMapel.query.filter_by(nip=guru.nip)
        
        # Ambil tahun unik dari tanggal ampu_mapel
        tahun_query = (
            db.session.query(extract('year', AmpuMapel.tanggal).label("tahun"))
            .filter_by(nip=guru.nip)
            .group_by(extract('year', AmpuMapel.tanggal))
            .order_by(extract('year', AmpuMapel.tanggal).desc())
            .all()
        )
    else:
        guru = Guru.query.all()
        data_ampu = AmpuMapel.query.all()
        daftar_pembagian = PembagianKelas.query.all()
        query = AmpuMapel.query

        # Ambil tahun unik dari tanggal ampu_mapel
        tahun_query = (
            db.session.query(extract('year', AmpuMapel.tanggal).label("tahun"))
            .group_by(extract('year', AmpuMapel.tanggal))
            .order_by(extract('year', AmpuMapel.tanggal).desc())
            .all()
        )
    # Ambil data untuk tampilan
    daftar_mapel = Mapel.query.all()
    daftar_semester = Semester.query.all()
    daftar_kelas = Kelas.query.all()
    daftar_tahun = TahunAkademik.query.order_by(TahunAkademik.id_tahun_akademik.desc()).all()
    
    tahun = request.args.get('tahun', type=int)
    bulan = request.args.get('bulan', type=int)
    tanggal = request.args.get('tanggal', type=int)
    page = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=10, type=int)
    # Mulai query dari AmpuMapel

    # Filter berdasarkan tanggal jika ada
    if tahun:
        query = query.filter(extract('year', AmpuMapel.tanggal) == tahun)
    if bulan:
        query = query.filter(extract('month', AmpuMapel.tanggal) == bulan)
    if tanggal:
        query = query.filter(extract('day', AmpuMapel.tanggal) == tanggal)

    # Urutkan berdasarkan ID (atau sesuaikan dengan kolom yang diinginkan)
    # Jangan urutkan pakai kolom dari table lain
    query = query.order_by(AmpuMapel.id_ampu.desc())


    # Pagination
    paginated_data = query.paginate(page=page, per_page=per_page, error_out=False)
    info_list = paginated_data.items
    # Cek total data dulu
    total_records = query.count()
    total_pages = (total_records + per_page - 1) // per_page

    # Jika halaman diminta melebihi total halaman, reset ke 1
    if page > total_pages:
        page = 1

    has_next = paginated_data.has_next
    has_prev = paginated_data.has_prev
    data_kbm = Kbm.query.all()
    page_range = range(max(1, page - 3), min(total_pages + 1, page + 3))
    keterangan = Keterangan.query.order_by(
        case(
            (Keterangan.id_keterangan == 1, 0),
            else_=1
        )
    ).all()

    # Serialize ke dict
    result = []
    for k in keterangan:
        result.append({
            "id_keterangan": k.id_keterangan,
            "keterangan": k.keterangan
        })

    for k in data_kbm:
        for l in info_list:
            if k.id_ampu == l.id_ampu:
                l.id_kbm = k.id_kbm
                l.materi = k.materi
                l.sub_materi = k.sub_materi
    # Ambil semua data pendukung untuk dropdown/filter
    data_guru = Guru.query.all()
    data_kelas = Kelas.query.all()
    data_mapel = Mapel.query.all()


    # Buat list tahun dari hasil query
    tahun_list = [int(row.tahun) for row in tahun_query if row.tahun is not None]
    return render_template("guru/list_ampu.html",
                            title="Pelajaran Diampu Anda",
                            btn_tambah=True,
                            thn=tahun_list,
                            data_guru = data_guru,
                            data_kelas = data_kelas,
                            data_mapel = data_mapel,
                            data_ampu=info_list,
                            data_kbm=data_kbm,
                            keterangan=result,
                            daftar_mapel=daftar_mapel,
                            daftar_semester=daftar_semester,
                            daftar_kelas=daftar_kelas,
                            daftar_tahun=daftar_tahun,
                            daftar_pembagian=daftar_pembagian,
                            page=page,
                            per_page=per_page,
                            total_pages=total_pages,
                            total_records=total_records,
                            has_next=has_next,
                            has_prev=has_prev,
                            page_range=page_range)


@app.route("/kbm/list/tambah", methods=["POST"])
def kbm_tambah():
    try:
        id_semester = request.json.get('id_semester')
        id_kelas = request.json.get('id_kelas')
        id_tahun_akademik = request.json.get('id_tahun_akademik')
        id_mapel = request.json.get("id_mapel", "").strip()
        tanggal_str = request.json.get("tanggal", "").strip()
        print(f"id_mapel: {id_mapel}, tanggal: {tanggal_str} id_semester: {id_semester}, id_kelas: {id_kelas}, id_tahun_akademik: {id_tahun_akademik}")
        if not Mapel.query.get(id_mapel):
            return jsonify({'msg': 'Mapel tidak ditemukan'}), 400

        # Konversi tanggal string ke objek datetime
        try:
            tanggal = datetime.strptime(tanggal_str, "%Y-%m-%d")
        except ValueError:
            print(f"Format tanggal tidak valid: {tanggal_str}")
            return jsonify({'msg': 'Format tanggal tidak valid (harus YYYY-MM-DD).'}), 400

        # Ambil nip tergantung role
        if session.get('role') == 'guru':
            nip = session.get('nip', '')
        else:
            nip = request.json.get('nip')
        print(f"nip: {nip}")
        # Validasi foreign key id_mapel
        if not Mapel.query.get(id_mapel):
            print(f"Mapel dengan ID {id_mapel} tidak ditemukan.")
            return jsonify({'msg': f'Mapel dengan ID "{id_mapel}" tidak ditemukan.'}), 400

        # Tambahkan AmpuMapel (biarkan id_ampu auto-increment)
        new_ampu = AmpuMapel(
            nip=nip,
            id_mapel=id_mapel,
            id_semester=id_semester,
            id_kelas=id_kelas,
            id_tahun_akademik=id_tahun_akademik,
            tanggal=tanggal
        )
        print(f"new_ampu: {new_ampu}")
        db.session.add(new_ampu)
        db.session.flush()  # agar new_ampu.id_ampu terisi
        print(new_ampu)
        # Tambahkan KBM
        new_kbm = Kbm(
            materi=request.json.get('materi'),
            sub_materi=request.json.get('sub_materi'),
            tanggal=tanggal_str,
            id_ampu=new_ampu.id_ampu
        )
        db.session.add(new_kbm)
        db.session.flush()  # agar new_kbm.id_kbm terisi

        # Ambil daftar siswa dan nama kelas
        siswa_kelas = PembagianKelas.query.filter_by(
            id_kelas=id_kelas,
            id_tahun_akademik=id_tahun_akademik
        ).all()
        nama_kelas = Kelas.query.get(id_kelas)
        print(f"nama_kelas: {nama_kelas}, siswa_kelas: {siswa_kelas}")
        if not nama_kelas:
            print(f"Kelas dengan ID {id_kelas} tidak ditemukan.")
            return jsonify({'msg': f'Kelas dengan ID "{id_kelas}" tidak ditemukan.'}), 400

        # Tambahkan kehadiran default "Hadir" (id_keterangan = 1)
        for siswa in siswa_kelas:
            exists = Kehadiran.query.filter_by(
                id_kbm=new_kbm.id_kbm,
                nis=siswa.nis,
                nama_kelas=nama_kelas.nama_kelas,
                id_keterangan=1
            ).first()
            if exists:
                continue

            obj = Kehadiran(
                id_kbm=new_kbm.id_kbm,
                nis=siswa.nis,
                nama_kelas=nama_kelas.nama_kelas,
                id_keterangan=None
            )
            db.session.add(obj)
            print(f"Menambahkan Kehadiran untuk NIS {siswa.nis} di kelas {nama_kelas.nama_kelas}")

        db.session.commit()
                
        print(f"KBM dan Kehadiran berhasil ditambahkan untuk kelas {nama_kelas.nama_kelas} dengan ID KBM {new_kbm.id_kbm}.")
        flash("Data berhasil ditambahkan", "success")
        return jsonify({'msg': 'Data berhasil ditambahkan'})

    except IntegrityError as e:
        db.session.rollback()
        return jsonify({'msg': 'Gagal menambahkan data karena konflik integritas.', 'error': str(e)}), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({'msg': 'Terjadi kesalahan internal.', 'error': str(e)}), 500

@app.route('/api/kehadiran/siswa')
def api_kehadiran_siswa():
    id_kbm = request.args.get('id_kbm')
    id_kelas = request.args.get('id_kelas')  # jika kamu pakai
    id_semester = request.args.get('id_semester')
    id_tahun = request.args.get('id_tahun_akademik')
    print(f"id_kbm: {id_kbm}, id_kelas: {id_kelas}, id_semester: {id_semester}, id_tahun: {id_tahun}")
    kehadiran_all = Kehadiran.query.filter_by(id_kbm=id_kbm).all()
    print(f"Jumlah kehadiran untuk id_kbm {id_kbm}: {len(kehadiran_all)}")
    pembagian_all = PembagianKelas.query.filter_by(id_kelas=id_kelas, id_tahun_akademik=id_tahun).all()
    print(f"Jumlah pembagian kelas untuk id_kelas {id_kelas} dan id_tahun_akademik {id_tahun}: {len(pembagian_all)}")
    siswa_list= []
    for i in kehadiran_all:
        for j in pembagian_all:
            print(f"Memeriksa NIS {i.nis} dengan {j.nis}")
            if i.nis == j.nis:
                siswa = {"nis": i.nis, "nama":i.siswa_rel.nama, "nama_kelas": i.nama_kelas,"id_kbm": i.id_kbm,"id_kehadiran":i.id_kehadiran , "id_keterangan": i.id_keterangan}
                siswa_list.append(siswa)
                print(f"Menambahkan nama kelas {i.nama_kelas} untuk NIS {i.nis}")
    
    print(siswa_list)
    return jsonify(siswa_list)
@app.route('/guru/kehadiran/edit', methods=["PUT"])
def edit_kehadiran_siswa():
    if request.is_json:
        data = request.get_json()
    else:
        # Fallback ke request.form jika bukan JSON
        data = request.form.to_dict(flat=False)  # biar bisa tampung multiple key
    
    print("DATA MASUK:", data)

    id_ampu = data.get('id_ampu', [None])[0] if isinstance(data.get('id_ampu'), list) else data.get('id_ampu')
    id_kbm = data.get('id_kbm', [None])[0] if isinstance(data.get('id_kbm'), list) else data.get('id_kbm')
    tanggal = time_zone_wib().date()

    print("KELAS:", data.get('nama_kelas'))

    # Loop semua key di request.form atau dict
    for key, value in data.items():
        if key.startswith("status["):
            nis = key[7:-1]  # ambil isi dalam kurung []
            print(nis)
            value= str(value).replace("['","").replace("']","")
            id_kehadiran = value.split("-")[1]
            print(id_kehadiran)
            value = value.split("-")[0]
            id_keterangan = int(value if not isinstance(value, list) else value[0])  # ambil statusnya
            print(id_keterangan)
            # cek apakah sudah ada record
            exists = Kehadiran.query.filter_by(id_kehadiran=id_kehadiran).first()

            if exists:
                exists.id_keterangan = id_keterangan

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"success": False, "message": "Gagal update kehadiran"}), 500

    return jsonify({"success": True, "message": "Kehadiran berhasil diupdate"}), 200

@app.route('/guru/kehadiran/tambah', methods=['POST'])
def simpan_kehadiran_siswa():
    id_ampu = request.form['id_ampu']
    id_kbm = request.form['id_kbm']
    tanggal = time_zone_wib().date()

    # Tambahkan data ke KBM (karena kehadiran harus ada ID KBM)
 
    print(request.form['nama_kelas'])
    for key in request.form:
        if key.startswith('status['):
            nis = key[7:-1]  # ambil angka di dalam status[20001]
            status = request.form[key]
            id_keterangan = status  # fungsi untuk mapping
            # Cek apakah sudah ada sebelumnya
            exists = Kehadiran.query.filter_by(
                id_kbm=id_kbm,
                nis=int(nis),
                nama_kelas=request.form['nama_kelas'],
                id_keterangan=id_keterangan
            ).first()

            if not exists:
                kehadiran = Kehadiran(
                    id_kbm=id_kbm,
                    nis=int(nis),
                    nama_kelas=request.form['nama_kelas'],
                    id_keterangan=id_keterangan
                )
                db.session.add(kehadiran)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()

    return '', 200
@app.route("/kbm/list/edit/<int:id>", methods=["PUT"])
def edit_ampu(id):
        ampu = AmpuMapel.query.get_or_404(id)
        ampu.id_mapel = request.json.get('id_mapel')
        ampu.id_pembagian = request.json.get('id_pembagian')
        ampu.id_semester = request.json.get('id_semester')
        ampu.id_tahun_akademik = request.json.get('id_tahun_akademik')
        db.session.commit()
        flash("Data berhasil diupdate", "success")
        
        return jsonify({'msg':'Data berhasil ditupdate'})
@app.route("/kbm/list/hapus/<int:id>", methods=["DELETE"])
def hapus_ampu(id):
    ampu = AmpuMapel.query.filter_by(id_ampu=id).first()
    if not ampu:
        return jsonify({'msg': 'Data ampu_mapel tidak ditemukan'}), 404

    # Hapus semua KBM yang terkait dengan ampu ini
    daftar_kbm = Kbm.query.filter_by(id_ampu=ampu.id_ampu).all()
    for kbm in daftar_kbm:
        # Hapus semua kehadiran yang terkait dengan ampu ini
        daftar_kehidaran = Kehadiran.query.filter_by(id_kbm=kbm.id_kbm).all()
        for kehadiran in daftar_kehidaran:
            db.session.delete(kehadiran)
        db.session.delete(kbm)
    db.session.delete(ampu)
    db.session.commit()

    flash("Data berhasil dihapus", "success")
    return jsonify({'msg': 'Data berhasil dihapus'})

@app.route('/guru/surat_izin')
def lihat_surat_izin_murid():

    siswa = Siswa.query.all()
    if not siswa:
        return "Siswa tidak ditemukan", 404

    data_kehadiran_all = Kehadiran.query.all()
    data_kehadiran = []
    for i in data_kehadiran_all:
        print(i.id_keterangan)
        print(i.surat_izin)
        if i.surat_izin != None :  # Hanya ambil yang keterangan 'Izin'
            print(f"Menambahkan kehadiran dengan id {i.id_kehadiran} karena keterangan 'Izin'")
            data_kehadiran.append(i)
        else:
            print(f"Melewati kehadiran dengan id {i.id_kehadiran} karena keterangan bukan 'Izin'")

    # Jika tidak ada data kehadiran, langsung render tanpa proses lanjut
    if not data_kehadiran:
        return render_template(
            'murid/kehadiran.html',
            data_kehadiran=None,
            btn_tambah=False,
            title="Kehadiran",
            title_data="Kehadiran"
        )

    pembagian = PembagianKelas.query.all()

    # Mapping pembagian berdasarkan tahun akademik
    pembagian_map = {
        p.id_tahun_akademik: {
            "nama_kelas": p.kelas_rel.nama_kelas,
            "tingkat": p.kelas_rel.tingkat
        } for p in pembagian
    }

    # Tambahkan info kelas ke data kehadiran
    enriched_kehadiran = []
    for d in data_kehadiran:
        nip_now = d.kbm_rel.ampu_rel.nip if d.kbm_rel else ''
        if nip_now == session.get('nip', ''):
            file_path = os.path.join(app.config['UPLOAD_SURAT_IZIN'],d.surat_izin if d.surat_izin else "")
            if not d.surat_izin or not os.path.exists(file_path):
                surat_izin = 'null'
            else:
                surat_izin = d.surat_izin
            enriched_kehadiran.append({
                "id_kehadiran": d.id_kehadiran,
                "surat_izin": surat_izin,
                "siswa_rel": d.siswa_rel,
                "kbm_rel": d.kbm_rel,
                "keterangan_rel": d.keterangan_rel,
                "nama_kelas": d.nama_kelas,
            })

    return render_template(
        'guru/surat_izin.html',
        data_kehadiran=enriched_kehadiran,
        btn_tambah=False,
        title="Surat Izin Murid",
        title_data="Surat Izin Murid"
    )
@app.route('/guru/penilaian')
def penilaian_list():
    if session.get('role') != 'guru':
        abort(403)

    nip = session.get('nip')
    if not nip:
        abort(403)

    # Ambil objek guru
    guru = Guru.query.filter_by(nip=nip).first()
    if not guru:
        abort(404)

    data_siswa = Siswa.query.all()
    data_ampu = AmpuMapel.query.filter_by(nip=nip).all()
    data_guru = Guru.query.all()
    data_kelas = Kelas.query.all()
    data_mapel = Mapel.query.all()

    title = "Nilai Siswa"
    title_data = "Nilai Siswa"
    btn_tambah = True

    # Ambil filter query
    tahun = request.args.get('tahun', type=int)
    bulan = request.args.get('bulan', type=int)
    tanggal = request.args.get('tanggal', type=int)
    page = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=10, type=int)
    nis = request.args.get('nis')
    id_kelas = request.args.get('id_kelas')
    id_mapel = request.args.get('id_mapel')
    jenis_penilaian = request.args.get('jenis_penilaian')
    query = (
        Penilaian.query
        .join(AmpuMapel, Penilaian.id_ampu == AmpuMapel.id_ampu)
        .filter(AmpuMapel.nip == nip)
    )

    if tahun:
        query = query.filter(extract('year', Penilaian.tanggal) == tahun)
    if bulan:
        query = query.filter(extract('month', Penilaian.tanggal) == bulan)
    if tanggal:
        query = query.filter(extract('day', Penilaian.tanggal) == tanggal)
    if nis:
        query = query.filter(Penilaian.nis == nis)
    if id_mapel:
        query = query.filter(AmpuMapel.id_mapel == id_mapel)
    if jenis_penilaian:
        query = query.filter(Penilaian.jenis_penilaian == jenis_penilaian)
    if id_kelas:
        subq = (
            db.session.query(PembagianKelas.nis)
            .filter(PembagianKelas.id_kelas == id_kelas)
            .subquery()
        )
        query = query.filter(Penilaian.nis.in_(subq))

    query = query.order_by(Penilaian.id_penilaian.desc())
    # Pagination
    paginated_data = query.paginate(page=page, per_page=per_page, error_out=False)
    info_list = paginated_data.items
    total_records = query.count()
    total_pages = paginated_data.pages
    has_next = paginated_data.has_next
    has_prev = paginated_data.has_prev
    page_range = range(max(1, page - 3), min(total_pages + 1, page + 3))

    # Ambil tahun unik dari tanggal AmpuMapel
    tahun_query = (
        db.session.query(extract('year', AmpuMapel.tanggal).label("tahun"))
        .filter(AmpuMapel.nip == nip)
        .group_by(extract('year', AmpuMapel.tanggal))
        .order_by(extract('year', AmpuMapel.tanggal).desc())
        .all()
    )
    thn = [int(row.tahun) for row in tahun_query if row.tahun is not None]

    return render_template( 'guru/penilaian.html', penilaian=info_list, tahun=thn, data_guru=data_guru, data_kelas=data_kelas, data_mapel=data_mapel, data_siswa=data_siswa, data_ampu=data_ampu, btn_tambah=btn_tambah, title=title, title_data=title_data, page=page, per_page=per_page, total_pages=total_pages, total_records=total_records, has_next=has_next, has_prev=has_prev, page_range=page_range )

@app.route('/guru/penilaian/tambah', methods=['POST'])
def tambah_penilaian():
    if session.get('role') != 'guru':
        abort(403)

    try:
        penilaian = Penilaian(
            nis=request.json.get('nis'),
            id_ampu=request.json.get('id_ampu'),
            jenis_penilaian=request.json.get('jenis_penilaian'),
            nilai=request.json.get('nilai'),
            tanggal=request.json.get('tanggal')  # Harus dalam format 'YYYY-MM-DD'
        )
        db.session.add(penilaian)
        db.session.commit()
        flash('Penilaian berhasil ditambahkan', 'success')
        return jsonify({'msg': 'Penilaian berhasil ditambahkan'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
@app.route('/guru/penilaian/tambah_excell', methods=['POST'])
def tambah_penilaian_excel():
    file = request.files.get('file')
    if not file:
        flash('❌ Tidak ada file yang diunggah.', 'danger')
        return redirect('/guru/penilaian')

    try:
        df = pd.read_excel(file)

        required_columns = ['nis', 'id_ampu', 'jenis_penilaian', 'nilai', 'tanggal']
        for col in required_columns:
            if col not in df.columns:
                flash(f'❌ Kolom "{col}" tidak ditemukan dalam file Excel.', 'danger')
                return redirect('/guru/penilaian')

        df = df[required_columns]

        # Drop baris yang kosong
        df.dropna(subset=required_columns, inplace=True)

        # Validasi dan konversi kolom
        failed_rows = []

        # Konversi NIS ke string
        try:
            df['nis'] = df['nis'].astype(int).astype(str)
        except Exception as e:
            flash(f'❌ Gagal konversi NIS: {e}', 'danger')
            return redirect('/guru/penilaian')

        # Nilai: Ganti koma jadi titik → float
        try:
            df['nilai'] = df['nilai'].astype(str).str.replace(',', '.').astype(float)
        except Exception as e:
            flash(f'❌ Gagal konversi nilai ke angka: {e}', 'danger')
            return redirect('/guru/penilaian')

        # Validasi nilai harus 0–100
        df = df[(df['nilai'] >= 0.0) & (df['nilai'] <= 100.0)]

        # Filter jenis_penilaian yang valid
        jenis_valid = ['UTS', 'UAS', 'UH', 'Tugas']
        df = df[df['jenis_penilaian'].isin(jenis_valid)]

        # Konversi tanggal ke datetime
        df['tanggal'] = pd.to_datetime(df['tanggal'], errors='coerce')
        df.dropna(subset=['tanggal'], inplace=True)

        inserted_count = 0

        for index, row in df.iterrows():
            row_index = index + 2  # baris di Excel
            error_msgs = []
            nis = str(row['nis'])
            id_ampu = row['id_ampu']
            if isinstance(id_ampu, float) and id_ampu.is_integer():
                    id_ampu = int(id_ampu)  # Convert float 123.0 to int 123 then to str => '123'
            else:
                    id_ampu = id_ampu
            print(id_ampu)
            siswa = Siswa.query.filter_by(nis=int(nis)).first()
            ampu = AmpuMapel.query.filter_by(id_ampu=id_ampu,nip=session.get('nip')).first()

            if not siswa:
                error_msgs.append(f"NIS tidak ditemukan")
            if not ampu:
                error_msgs.append(f"id_ampu tidak ditemukan")

            if pd.isna(row['tanggal']):
                error_msgs.append("Tanggal tidak valid")

            if error_msgs:
                failed_rows.append({
                    'row': row_index,
                    'nis': nis,
                    'nama': siswa.nama if siswa else '-',
                    'error': error_msgs
                })
                continue

            try:
                penilaian = Penilaian(
                    nis=nis,
                    id_ampu=id_ampu,
                    jenis_penilaian=row['jenis_penilaian'],
                    nilai=row['nilai'],
                    tanggal=row['tanggal'].date()
                )
                db.session.add(penilaian)
                inserted_count += 1

            except Exception as db_error:
                failed_rows.append({
                    'row': row_index,
                    'nis': nis,
                    'nama': siswa.nama if siswa else '-',
                    'error': [f'Gagal simpan ke database: {str(db_error)}']
                })
                db.session.rollback()
                continue

        db.session.commit()

        flash(f'✅ {inserted_count} penilaian berhasil ditambahkan.', 'success')

        if failed_rows:
            for fail in failed_rows:
                flash(f"❌ Baris {fail['row']} | {fail['nama']} (NIS: {fail['nis']}) gagal: {', '.join(fail['error'])}", 'danger')

    except Exception as e:
        db.session.rollback()
        flash(f'❌ Gagal memproses file: {str(e)}', 'danger')

    return redirect('/guru/penilaian')

@app.route('/guru/penilaian/edit/<int:id_penilaian_old>', methods=['PUT'])
def edit_penilaian(id_penilaian_old):
    if session.get('role') != 'guru':
        abort(403)

    penilaian = Penilaian.query.get(id_penilaian_old)
    if not penilaian:
        return jsonify({'error': 'Penilaian tidak ditemukan'}), 404

    try:
        penilaian.nis = request.json.get('nis', penilaian.nis)
        penilaian.id_ampu = request.json.get('id_ampu', penilaian.id_ampu)
        penilaian.jenis_penilaian = request.json.get('jenis_penilaian', penilaian.jenis_penilaian)
        penilaian.nilai = request.json.get('nilai', penilaian.nilai)
        tgl_str = request.json.get('tanggal')
        if tgl_str:
            penilaian.tanggal = datetime.strptime(tgl_str, "%Y-%m-%d") # Format: 'YYYY-MM-DD'

        db.session.commit()
        flash('Penilaian berhasil diperbarui', 'info')
        return jsonify({'msg': 'Penilaian berhasil diperbarui'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/guru/penilaian/hapus/<id_penilaian>', methods=['DELETE'])
def hapus_penilaian(id_penilaian):
    if session.get('role') != 'guru':
        abort(403)
    penilaian = Penilaian.query.filter_by(id_penilaian=id_penilaian).first()
    if not penilaian:
        return jsonify({'error': 'penilaian tidak ditemukan'}), 404
    db.session.delete(penilaian)
    db.session.commit()
    flash('penilaian berhasil dihapus', 'warning')
    return jsonify({'msg': 'penilaian berhasil dihapus'})

@app.route('/guru/pengumuman')
def view_guru_pengumuman():
    berita = Berita.query.filter_by(pengumuman_untuk='guru').all()
    btn_tambah = False
    title = "Pengumuman"
    title_data = "Pengumuman"
    return render_template('guru/berita.html', berita=berita,btn_tambah=btn_tambah,title=title,title_data=title_data)
@app.route('/guru/evaluasi_guru')
def evaluasi_guru():
    if session.get('role') != 'guru':
        abort(403)

    nip = session.get('nip')
    guru = Guru.query.filter_by(nip=nip).first_or_404()

    # Ambil daftar tahun akademik & semester master (untuk filter dropdown kalau perlu)
    tahun_list = TahunAkademik.query.order_by(TahunAkademik.mulai.desc()).all()
    semester_master = Semester.query.order_by(Semester.id_semester.asc()).all()

    # Tahun & semester aktif (default)
    tahun_semester_aktif = TahunSemester.query.order_by(TahunSemester.mulai.desc()).first()
    tahun_id = request.args.get('tahun_id', type=int)
    semester_id = request.args.get('semester_id')
    print(tahun_id)
    print(semester_id)

    print(tahun_id)
    print(semester_id)
    # Evaluasi detail untuk guru tsb (filter by tahun_id & semester_id kalau user pilih)
    evaluasi_list = (
        EvaluasiGuru.query
        .filter_by(nip=nip)
        .all()
    )
    
    # Ambil semua tahunsemester (dari lama ke baru → perjalanan)
    query = TahunSemester.query

    # Filter hanya kalau ada input dari request
    if tahun_id:
        query = query.filter(TahunSemester.tahun_id == tahun_id)
    if semester_id:
        query = query.filter(TahunSemester.semester_id == semester_id)

    tahun_semester_list = query.order_by(TahunSemester.mulai.asc()).all()

    chart_data = []
    for ts in tahun_semester_list:
        rata = (
            db.session.query(
                func.avg(
                    (EvaluasiGuru.q1 + EvaluasiGuru.q2 + EvaluasiGuru.q3 +
                     EvaluasiGuru.q4 + EvaluasiGuru.q5 + EvaluasiGuru.q6 +
                     EvaluasiGuru.q7 + EvaluasiGuru.q8 + EvaluasiGuru.q9 +
                     EvaluasiGuru.q10 + EvaluasiGuru.q11) / 11.0
                )
            )
            .filter(
                EvaluasiGuru.nip == nip,
                EvaluasiGuru.tahun_id == ts.tahun_id,
                EvaluasiGuru.semester_id == ts.semester_id
            )
            .scalar()
        )
        chart_data.append({
            "label": f"{ts.tahun.tahun_akademik } - {ts.semester.semester}",  # contoh: "2024/2025 - Ganjil"
            "rata": round(rata, 2) if rata else 0
        })

    return render_template( "guru/evaluasi_guru.html", tahun_aktif=tahun_id, semester_aktif=semester_id, tahun_list=tahun_list, semester_list=semester_master, evaluasi=evaluasi_list, chart_data=chart_data, guru=guru, btn_tambah=False, title="Manage Evaluasi Guru", title_data="Evaluasi Guru" )

def _get_filter_params():
    tahun_id = request.args.get('tahun_id', type=int)
    semester_id = request.args.get('semester_id')  # CHAR(1) → biarkan string
    return tahun_id, semester_id

def _get_tahun_semester_aktif(tahun_id, semester_id):
    if not (tahun_id and semester_id):
        return None
    # SESUAIKAN nama kolom di model TahunSemester: tahun_id/semester_id atau id_tahun/id_semester
    return TahunSemester.query.filter_by(tahun_id=tahun_id, semester_id=semester_id).first()
# Ekspresi nilai rata-rata
def _score_expr():
    return (
        (EvaluasiGuru.q1 + EvaluasiGuru.q2 + EvaluasiGuru.q3 + EvaluasiGuru.q4 + EvaluasiGuru.q5 +
         EvaluasiGuru.q6 + EvaluasiGuru.q7 + EvaluasiGuru.q8 + EvaluasiGuru.q9 + EvaluasiGuru.q10 + EvaluasiGuru.q11)
        / literal(11.0)
    )
def _apply_scope_filters(query):
    role = session.get('role')
    if role not in ('admin', 'guru'):
        abort(403)

    tahun_id, semester_id = _get_filter_params()

    if role == 'guru':
        nip = session.get('nip')
        query = query.filter(EvaluasiGuru.nip == nip)

    if tahun_id:
        query = query.filter(EvaluasiGuru.tahun_id == tahun_id)
    if semester_id:
        query = query.filter(EvaluasiGuru.semester_id == semester_id)

    return query, role

def _summary_query():
    score = _score_expr()
    q = (
        db.session.query(
            EvaluasiGuru.nip.label('nip'),
            Guru.nama.label('nama_guru'),
            TahunAkademik.tahun_akademik.label('tahun'),
            Semester.semester.label('semester'),
            func.avg(score).label('rata'),
            func.count(EvaluasiGuru.nip).label('jml_respon')
        )
        .join(Guru, Guru.nip == EvaluasiGuru.nip)
        .join(TahunAkademik, TahunAkademik.id_tahun_akademik == EvaluasiGuru.tahun_id)
        .join(Semester, Semester.id_semester == EvaluasiGuru.semester_id)
    )
    q, role = _apply_scope_filters(q)
    q = q.group_by('nip', 'nama_guru', 'tahun', 'semester') \
         .order_by(Guru.nama.asc(), TahunAkademik.mulai.asc(), Semester.id_semester.asc())
    return q, role

def _detail_query():
    score = _score_expr()
    q = (
        db.session.query(
            EvaluasiGuru.nip.label('nip'),
            Guru.nama.label('nama_guru'),
            TahunAkademik.tahun_akademik.label('tahun'),
            Semester.semester.label('semester'),
            EvaluasiGuru.q1, EvaluasiGuru.q2, EvaluasiGuru.q3, EvaluasiGuru.q4, EvaluasiGuru.q5,
            EvaluasiGuru.q6, EvaluasiGuru.q7, EvaluasiGuru.q8, EvaluasiGuru.q9, EvaluasiGuru.q10, EvaluasiGuru.q11,
            score.label('rata')
        )
        .join(Guru, Guru.nip == EvaluasiGuru.nip)
        .join(TahunAkademik, TahunAkademik.id_tahun_akademik == EvaluasiGuru.tahun_id)
        .join(Semester, Semester.id_semester == EvaluasiGuru.semester_id)
    )
    q, role = _apply_scope_filters(q)
    q = q.order_by(Guru.nama.asc(), TahunAkademik.mulai.asc(), Semester.id_semester.asc())
    return q, role
@app.route('/evaluasi_guru/pdf')
def evaluasi_guru_pdf():
    q, role = _summary_query()
    rows = q.all()

    tahun_id, semester_id = _get_filter_params()
    ts_aktif = _get_tahun_semester_aktif(tahun_id, semester_id)

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    judul = "Laporan Evaluasi Guru - Ringkasan"
    scope = f"Role: {role}"
    periode = (
        f"Periode: {ts_aktif.tahun.tahun_akademik} - {ts_aktif.semester.semester}"
        if ts_aktif else "Periode: Semua Tahun & Semester"
    )
    story.append(Paragraph(judul, styles['Title']))
    story.append(Paragraph(scope, styles['Normal']))
    story.append(Paragraph(periode, styles['Normal']))
    story.append(Spacer(1, 0.3 * cm))

    data = [["NIP", "Nama Guru", "Tahun", "Semester", "Rata-rata", "Jml Respon"]]
    if rows:
        for r in rows:
            data.append([
                r.nip,
                r.nama_guru or "-",
                r.tahun or "-",
                r.semester or "-",
                f"{(r.rata or 0):.2f}",
                int(r.jml_respon or 0)
            ])
    else:
        data.append(["-", "Tidak ada data", "-", "-", "-", "-"])

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#333333')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#999999')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"evaluasi_guru_{role}_{tahun_id or 'all'}_{semester_id or 'all'}_{ts}.pdf"
    return send_file(buf, mimetype='application/pdf', download_name=filename, as_attachment=True)
# Alias: kamu minta /excell, aku buat dua-duanya /excell dan /excel
@app.route('/evaluasi_guru/excell')
@app.route('/evaluasi_guru/excel')
def evaluasi_guru_excel():
    q_sum, role = _summary_query()
    sum_rows = q_sum.all()

    q_det, _ = _detail_query()
    det_rows = q_det.all()

    tahun_id, semester_id = _get_filter_params()
    ts_aktif = _get_tahun_semester_aktif(tahun_id, semester_id)
    periode_txt = (
        f"{ts_aktif.tahun.tahun_akademik} - {ts_aktif.semester.semester}"
        if ts_aktif else "Semua Tahun & Semester"
    )

    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F0F0F0")

    # Sheet Ringkasan
    ws = wb.active
    ws.title = "Ringkasan"

    # Title + Periode
    ws.merge_cells('A1:F1')
    ws['A1'] = "Laporan Evaluasi Guru - Ringkasan"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Role: {role} | Periode: {periode_txt}"
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])  # baris 3 kosong

    headers_sum = ["NIP", "Nama Guru", "Tahun", "Semester", "Rata-rata", "Jml Respon"]
    ws.append(headers_sum)  # baris 4

    # Styling header
    for c in range(1, len(headers_sum) + 1):
        cell = ws.cell(row=4, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = border_all

    # Data mulai baris 5
    start_row = 5
    if sum_rows:
        for idx, r in enumerate(sum_rows):
            row = [
                r.nip,
                r.nama_guru or "",
                r.tahun or "",
                r.semester or "",
                float(f"{(r.rata or 0):.2f}"),
                int(r.jml_respon or 0),
            ]
            ws.append(row)
            for c in range(1, len(headers_sum) + 1):
                cell = ws.cell(row=start_row + idx, column=c)
                cell.border = border_all
                cell.alignment = Alignment(horizontal="center" if c != 5 else "right")
    else:
        ws.append(["-", "Tidak ada data", "-", "-", 0.0, 0])
        for c in range(1, len(headers_sum) + 1):
            cell = ws.cell(row=start_row, column=c)
            cell.border = border_all
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.freeze_panes = "A5"

    # Sheet Detail
    ws2 = wb.create_sheet(title="Detail")
    ws2.merge_cells('A1:P1')
    ws2['A1'] = "Laporan Evaluasi Guru - Detail"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.merge_cells('A2:P2')
    ws2['A2'] = f"Role: {role} | Periode: {periode_txt}"
    ws2['A2'].alignment = Alignment(horizontal='center')

    ws2.append([])  # baris 3 kosong

    headers_det = ["NIP", "Nama Guru", "Tahun", "Semester",
                   "Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10","Q11","Rata-rata"]
    ws2.append(headers_det)  # baris 4

    for c in range(1, len(headers_det) + 1):
        cell = ws2.cell(row=4, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = border_all

    start_row2 = 5
    if det_rows:
        for idx, r in enumerate(det_rows):
            row = [
                r.nip,
                r.nama_guru or "",
                r.tahun or "",
                r.semester or "",
                r.q1, r.q2, r.q3, r.q4, r.q5, r.q6, r.q7, r.q8, r.q9, r.q10, r.q11,
                float(f"{(r.rata or 0):.2f}")
            ]
            ws2.append(row)
            for c in range(1, len(headers_det) + 1):
                cell = ws2.cell(row=start_row2 + idx, column=c)
                cell.border = border_all
                cell.alignment = Alignment(horizontal="center")
    else:
        ws2.append(["-", "Tidak ada data", "-", "-"] + [0]*12)
        for c in range(1, len(headers_det) + 1):
            cell = ws2.cell(row=start_row2, column=c)
            cell.border = border_all
            cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    for col in "EFGHIJKLMNOP":
        ws2.column_dimensions[col].width = 10
    ws2.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"evaluasi_guru_{role}_{tahun_id or 'all'}_{semester_id or 'all'}_{ts}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename, as_attachment=True)